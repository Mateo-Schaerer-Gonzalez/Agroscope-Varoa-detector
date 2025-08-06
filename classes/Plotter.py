import os
import matplotlib
matplotlib.use('Agg')  # Must be set before importing pyplot



import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import pandas as pd
import cv2
import math
from classes.mite import Mite
from openpyxl.drawing.image import Image as XLImage


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class Plotter:
    def __init__(self, stage, output_folder, discobox_run):

        self.stage = stage
        base_dir = os.path.dirname(os.path.abspath(__file__))

        # set the outputfolder
        if discobox_run:
            # get the output path
            self.output_path = os.path.abspath(os.path.join(base_dir, "..",  "..", output_folder))
        else:
            self.output_path = os.path.abspath(os.path.join(base_dir,"..",  output_folder))

        general_summary_path = os.path.abspath(os.path.join(self.output_path, os.pardir))

        self.pdf_path = os.path.join(self.output_path, "recording.pdf")
        self.csv_path = os.path.join(self.output_path, "summary.csv")
        self.frame_path = os.path.join(self.output_path, "frame_0.jpg")
        self.survival_path = os.path.join(self.output_path, "survival.png")
        self.time_survival_path = os.path.join(general_summary_path, "surivival.png")
        self.distribution_max_diff = os.path.join(general_summary_path, 'max_diff.png')
        self.distribution_max_local_diff = os.path.join(general_summary_path,'local_diff.png')
        self.distribution_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'variabilites.csv')
        self.by_mite_path = os.path.join(general_summary_path,'zones')
        self.excel_by_zones = os.path.join(self.by_mite_path, "mites.xlsx")
        self.excel_by_recording = os.path.join(general_summary_path, "recordings_summary.xlsx")
        os.makedirs(self.by_mite_path, exist_ok=True)

    
    def make_survival_graph(self, recording_number):
        summary_data = self.stage.data
        mite_data = self.stage.mite_data


        # Filter both datasets by the given recording number
        summary_data = summary_data[summary_data['recording'] == recording_number]
        mite_data = mite_data[mite_data['recording'] == recording_number]

        # Create a figure with 2 subplots side by side
        fig, axes = plt.subplots(1, 2, figsize=(12, 6))

        # Subplot 1: Histogram of max pixel difference
        axes[0].hist(mite_data['max diff'], bins=20, color="steelblue", edgecolor="black")
        axes[0].set_title("Distribution of Mite Max Pixel Difference")
        axes[0].set_xlabel("Max Pixel Difference")
        axes[0].set_ylabel("Frequency")
        axes[0].grid(True)
        axes[0].axvline(Mite.threshold, color='red', linestyle='--', linewidth=2)

        # Subplot 2: Bar chart of survival rates
        axes[1].bar(summary_data['Zone ID'], summary_data['Survival %'], color="mediumseagreen", edgecolor="black")
        axes[1].set_title("Survival Rate by Zone")
        axes[1].set_xlabel("Zone ID")
        axes[1].set_ylabel("Survival %")
        axes[1].set_ylim(0, 100)
        axes[1].tick_params(axis='x', rotation=45)

        plt.tight_layout()
        plt.savefig(self.survival_path)
        plt.close()

    
    def create_recording_pdf(self, recording_count):
        
        with PdfPages(self.pdf_path) as pdf:
            # A4 size in inches: 8.27 x 11.69
            fig = plt.figure(figsize=(8.27, 11.69))

           
            n_rows = 3
            gs = fig.add_gridspec(n_rows, 1, height_ratios=[1, 3, 1.5])

            # --- Section 1: Summary Table ---
            
            df = self.stage.data
            df = df[df['recording'] == recording_count]
            print(df.head())
            ax1 = fig.add_subplot(gs[0])
            ax1.axis('off')
            table = ax1.table(cellText=df.values, colLabels=df.columns, loc='center', cellLoc='center')
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.0)
            ax1.set_title("Zone Survival Summary", fontsize=12, pad=10)

            # --- Section 2: Frame Image ---
            if os.path.exists(self.frame_path):
                img = plt.imread(self.frame_path)
                ax2 = fig.add_subplot(gs[1])
                ax2.imshow(img)
                ax2.axis('off')
                ax2.set_title("Detection output", fontsize=12, pad=10)

            # --- Section 3: Survival Plot ---
            if os.path.exists(self.survival_path):
                img2 = plt.imread(self.survival_path)
                ax3 = fig.add_subplot(gs[2])
                ax3.imshow(img2)
                ax3.axis('off')
                ax3.set_title("Survival Rate by Zone", fontsize=12, pad=10)

            plt.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

        print(f"PDF successfully saved to: {self.pdf_path}")


    def make_survival_time_graph(self):
        df = self.stage.data
        plt.figure(figsize=self.stage.img_size)

        plt.figure(figsize=(10, 6))

        for zone in df['Zone ID'].unique():
            zone_data = df[df['Zone ID'] == zone]
            plt.plot(
                zone_data['recording'],
                zone_data['Survival %'],
                marker='o',
                label=f'{zone}'
            )

        plt.xlabel('Recording Count')
        plt.ylabel('Survival %')
        plt.title('Survival Percentage Over Recordings Per Zone')
        plt.legend(title='Zone ID')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(self.time_survival_path)
        plt.close()



    def distribution_graph(self):
        df = pd.read_csv(self.distribution_data_path)

        # Plot histogram for each group (Alive True/False)
        plt.figure(figsize=self.stage.img_size)

        for alive_status, group in df.groupby('Ground_truth'):
            plt.hist(group['max_diff'], bins=30, alpha=0.6, label=alive_status)

        
        plt.xlabel('max difference')
        plt.ylabel('Frequency')
        plt.title('Histogram of local difference')
        plt.legend()
        plt.savefig(self.distribution_max_diff)
        plt.close()


        for alive_status, group in df.groupby('Ground_truth'):
            plt.hist(group['local_diff'], bins=30, alpha=0.6, label=alive_status)


        plt.xlabel('local difference')
        plt.ylabel('Frequency')
        plt.title('Histogram of local difference')
        plt.legend()
        plt.savefig(self.distribution_max_local_diff)
        plt.close()

    def save_frame0_detection(self, image, thickness=2):
        # Draw zones on the image
        for zone in self.stage.zones:
            zone.draw(image, thickness=thickness)

        # Save image with a unique name
        filename = os.path.join(self.output_path, "frame_0.jpg")

        cv2.imwrite(filename, image)
        print(f"Image saved to: {filename}")

    def plot_variability_by_mite(self):

        # Ensure data is sorted properly
        df = self.stage.mite_data.sort_values(by=['zone ID', 'mite ID', 'recording'])

        # Create zones folder if not already
        os.makedirs(os.path.join(self.by_mite_path, "zones"), exist_ok=True)

        # Set threshold and consistent y-limits
        threshold = Mite.threshold
        y_min = df[['max diff', 'local diff']].min().min() - 0.1
        y_max = df[['max diff', 'local diff']].max().max() + 0.1

        # Group by zone
        zones = df['zone ID'].unique()

        for zone in zones:
            path = os.path.join(self.by_mite_path, "zones", f'{zone}.png')

            zone_df = df[df['zone ID'] == zone]
            mites = zone_df['mite ID'].unique()
            num_mites = len(mites)

            # Arrange in two columns
            ncols = 2
            nrows = math.ceil(num_mites / ncols)

            fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(14, 4 * nrows), sharex=True, sharey=True)
            axes = axes.flatten()

            fig.suptitle(f"Zone: {zone}", fontsize=18)

            for i, mite_id in enumerate(mites):
                ax = axes[i]
                mite_df = zone_df[zone_df['mite ID'] == mite_id]

                # Background shading
                ax.axhspan(y_min, threshold, color='red', alpha=0.2)
                ax.axhspan(threshold, y_max, color='green', alpha=0.2)

                # Plot data
                ax.plot(mite_df['recording'], mite_df['max diff'], marker='o', label='Max Diff')
                ax.plot(mite_df['recording'], mite_df['local diff'], marker='x', label='Local Diff')

              

                ax.set_title(f"Mite: {mite_id}")
                ax.set_ylabel("Diff")
                ax.set_ylim(y_min, y_max)
                ax.grid(True)
                ax.legend()
            print(f"plotting zone: {zone}")

            # Hide unused subplots
            for j in range(len(mites), len(axes)):
                fig.delaxes(axes[j])

            axes[-1].set_xlabel("Recording")
            plt.tight_layout(rect=[0, 0.03, 1, 0.95])
            plt.savefig(path)
            plt.close()

    def excel_summary_recordings(self):
        summary_df = self.stage.data
        wb = Workbook()
        wb.remove(wb.active)

        recordings = sorted(summary_df['recording'].unique())

        # add overview:
        ws = wb.create_sheet(title="Overview")
        try:
            img = XLImage(self.time_survival_path)
            img.anchor = 'A1'  # position on overview sheet, adjust as needed
            ws.add_image(img)
        except Exception as e:
            print(f"Failed to add overview image: {e}")

        for rec in recordings:
            rec_df = summary_df[summary_df['recording'] == rec]

            ws = wb.create_sheet(title=f"Recording {rec}")

            for r_idx, row in enumerate(dataframe_to_rows(rec_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # add image
            try:
                path = os.path.join(os.path.dirname(self.excel_by_recording), f'recording{rec}', 'survival.png' )
                print(path)
                img = XLImage(path)
                img.anchor = 'H1'
                ws.add_image(img)

            except Exception as e:
                print("adding image failed")

        path = self.excel_by_recording  # define this path in your class
        wb.save(path)
        print(f"Summary Excel with sheets per recording saved to {path}")

    

    def excel_summary_mites(self):
         # Define color fills
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red

        df = self.stage.mite_data

        # Get unique zones
        zones = df['zone ID'].unique()

        # Create a new Excel workbook
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        for zone in zones:
            zone_df = df[df['zone ID'] == zone]
            
            # Pivot so rows = mite ID, cols = recording, values = status
            pivot = zone_df.pivot(index='mite ID', columns='recording', values='status')

            # Add a new worksheet for the zone
            ws = wb.create_sheet(title=str(zone))

            # Write headers and data
            for r_idx, row in enumerate(dataframe_to_rows(pivot.reset_index(), index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx > 1 and c_idx > 1:  # Only data cells (skip headers)
                        if value == 'alive':
                            cell.fill = green_fill
                        elif value == 'dead':
                            cell.fill = red_fill

            img_path = os.path.join(self.by_mite_path, "zones", f"{zone}.png")
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.width = img.width * 0.5
                img.height = img.height * 0.5

                img.anchor = "H2"  # Adjust as needed (e.g., "H2" puts image starting at column H row 2)
                ws.add_image(img)
            else:
                print(f"Warning: Image for zone '{zone}' not found at {img_path}")

        
        wb.save(self.excel_by_zones)
        print(f"Excel summary with zone-wise sheets saved to {self.excel_by_zones}")
            


